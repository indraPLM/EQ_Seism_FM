# -*- coding: utf-8 -*-
"""
Streamlit App: BMKG Focal Mechanism Viewer with Summary and Beachball Export
Created by: Indra Gunawan
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from obspy.imaging.beachball import beach
from bs4 import BeautifulSoup
from io import BytesIO
import base64
import folium
from streamlit_folium import st_folium

# 🌍 Page config
st.set_page_config(page_title="BMKG Focal Viewer", layout="wide", page_icon="🌋")

# 🛠 Sidebar Inputs
st.sidebar.header("BMKG Focal Filter")
start_time = st.sidebar.text_input("Start Time", "2025-06-01 00:00:00")
end_time = st.sidebar.text_input("End Time", "2025-06-30 23:59:59")
col1, col2 = st.sidebar.columns(2)
North = float(col1.text_input("North", "6.0"))
South = float(col2.text_input("South", "-13.0"))
col3, col4 = st.sidebar.columns(2)
West = float(col3.text_input("West", "90.0"))
East = float(col4.text_input("East", "142.0"))

st.sidebar.header("Global CMT Filter :")
cmt_start=st.sidebar.text_input('Start Time:', '2015-01-01 00:00')
cmt_end=st.sidebar.text_input('End Time:', '2020-12-31 23:59')

# 📦 Fetch BMKG catalog
@st.cache_data
def load_bmkg_focal(url):
    res = requests.get(url)
    raw_text = res.text.strip()
    lines = raw_text.split("\n")
    rows = [line.split("|") for line in lines if "|" in line]
    return rows


url = "http://202.90.198.41/qc_focal.txt"
rows = load_bmkg_focal(url)

base_cols = ['date_time', 'mode', 'status', 'phase', 'mag', 'type_mag','count','azgap','RMS',
             'lat', 'lon', 'depth', 'S1', 'D1', 'R1', 'S2', 'D2', 'R2','Fit','DC','CLVD','type','location']
n_extra = max(0, len(rows[0]) - len(base_cols)) if rows else 0
cols = base_cols + [f'extra_{i}' for i in range(n_extra)]
df = pd.DataFrame(rows[1:], columns=cols)

# 🔁 Preprocess columns
def fix_coord(val, axis):
    val = val.strip()
    return -float(val.strip('S')) if val.endswith('S') else float(val.strip('N')) if axis == 'lat' else \
           -float(val.strip('W')) if val.endswith('W') else float(val.strip('E'))

def fix_float(col): return pd.to_numeric(col, errors='coerce')

df['fixedLat'] = df['lat'].apply(lambda x: fix_coord(x, 'lat'))
df['fixedLon'] = df['lon'].apply(lambda x: fix_coord(x, 'lon'))
df['date_time'] = pd.to_datetime(df['date_time'], errors='coerce')
for col in ['mag','depth','S1','D1','R1','S2','D2','R2']: df[col] = fix_float(df[col])

df = df[
    (df['date_time'] >= start_time) & (df['date_time'] <= end_time) &
    (df['fixedLat'].between(South, North)) &
    (df['fixedLon'].between(West, East))
]

def get_beachball_width(east, west):
    dist_lon = abs(east - west)
    if dist_lon > 55:
        return 1.5
    elif 40 < dist_lon <= 55:
        return 1.3
    elif 30 < dist_lon <= 40:
        return 1.1
    elif 15 < dist_lon <= 30:
        return 0.9
    elif 10 < dist_lon <= 15:
        return 0.7
    elif 5 < dist_lon <= 10:
        return 0.5
    else:
        return 0.3

w = get_beachball_width(East, West)

# 🗺️ Cartopy Plot
st.markdown(f"### 🗺️ Peta Focal Mechanism BMKG \n{start_time} – {end_time}")

fig = plt.figure(dpi=300)
ax = fig.add_subplot(111, projection=ccrs.PlateCarree(central_longitude=120))
ax.set_extent((West, East, South-0.5, North+0.5))
ax.add_feature(cfeature.BORDERS, linestyle='-', linewidth=0.5, alpha=0.5)
ax.coastlines(resolution='10m', color='black', linewidth=0.5, alpha=0.5)
for _, row in df.iterrows():
    if pd.notnull(row["S1"]) and pd.notnull(row["D1"]) and pd.notnull(row["R1"]):
        x, y = ax.projection.transform_point(row["fixedLon"], row["fixedLat"], ccrs.Geodetic())
        color = "r" if row["depth"] < 60 else "y" if row["depth"] < 300 else "g"
        ball = beach([row["S1"], row["D1"], row["R1"]],
             xy=(x, y), width=w, linewidth=0.5,
             alpha=0.65, zorder=10, facecolor=color)

        ax.add_collection(ball)
st.pyplot(fig)

summary_df = df[['date_time', 'mag', 'type_mag', 'fixedLat', 'fixedLon', 'depth',
                 'S1', 'D1', 'R1', 'S2', 'D2', 'R2', 'location']].copy()

# Rename columns for clarity
summary_df.columns = ['DateTime', 'Magnitude', 'Type Magnitude', 'Latitude', 'Longitude', 'Depth',
                      'Strike NP1', 'Dip NP1', 'Rake NP1', 'Strike NP2', 'Dip NP2', 'Rake NP2', 'Remark']
summary_df.index = range(1, len(summary_df) + 1)  # Reindex starting from 1
st.dataframe(summary_df)

# Optional: Add beachball mechanism visuals or export column (e.g., if exporting plots)

# Reuse or clone the summary_df you already created earlier
report_df = summary_df.copy()
report_df.index = range(len(report_df))  # match image naming convention

from obspy.imaging.beachball import beachball
import matplotlib.pyplot as plt

def get_color(depth):
    return 'r' if depth <= 60 else 'yellow' if depth <= 300 else 'g'

def generate_beachball_images(df, prefix="cmt"):
    filenames = []
    for idx, row in df.iterrows():
        mt = [row['Strike NP1'], row['Dip NP1'], row['Rake NP1']]
        color = get_color(row['Depth'])
        fig = beachball(mt, facecolor=color)
        filename = f"{prefix}_{idx}.png"
        fig.savefig(filename)
        plt.close(fig.figure)
        filenames.append(filename)
    return filenames

report_df['Focal'] = generate_beachball_images(report_df)

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

def export_excel_with_images(df, filename="focal_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Earthquake Focals"
    
    headers = list(df.columns.drop('Focal')) + ['Focal']
    ws.append(headers)

    for i, row in df.iterrows():
        ws.append(list(row.drop('Focal')) + [''])  # image goes into last column
        img = XLImage(row['Focal'])
        img.height = 50
        img.width = 50
        cell = f"{chr(65 + len(headers))}{i + 2}"  # Excel position
        ws.add_image(img, cell)

    wb.save(filename)

def image_to_html(path):
    return f'<img src="{path}" width="60">'

html_file = "focal_report.html"
report_df.to_html(html_file, escape=False, formatters={'Focal': image_to_html})

from fpdf import FPDF
from PIL import Image

def export_to_pdf(df, filename="focal_report.pdf"):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    # Header
    pdf.set_fill_color(200, 220, 255)
    pdf.cell(0, 10, "Earthquake Focal Mechanism Summary", ln=True, fill=True)

    # Table Headers
    headers = ['DateTime', 'Magnitude', 'Depth', 'Strike', 'Dip', 'Rake', 'Image']
    col_widths = [40, 20, 20, 20, 20, 20, 30]
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 10, h, border=1)
    pdf.ln()

    for _, row in df.iterrows():
        pdf.cell(col_widths[0], 10, str(row['DateTime']), border=1)
        pdf.cell(col_widths[1], 10, f"{row['Magnitude']:.1f}", border=1)
        pdf.cell(col_widths[2], 10, f"{row['Depth']:.1f}", border=1)
        pdf.cell(col_widths[3], 10, f"{row['Strike NP1']:.1f}", border=1)
        pdf.cell(col_widths[4], 10, f"{row['Dip NP1']:.1f}", border=1)
        pdf.cell(col_widths[5], 10, f"{row['Rake NP1']:.1f}", border=1)

        # Resize image for display
        img_path = row['Focal']
        img = Image.open(img_path)
        img.thumbnail((25, 25))
        temp_path = f"thumb_{img_path}"
        img.save(temp_path)

        x = pdf.get_x()
        y = pdf.get_y()
        pdf.cell(col_widths[6], 10, '', border=1)
        pdf.image(temp_path, x + 1, y + 1, h=8)
        pdf.ln()

    pdf.output(filename)

# Generate the PDF
export_to_pdf(report_df)

# Show download button
with open("focal_report.pdf", "rb") as f:
    st.download_button("⬇️ Download PDF Report", f.read(), file_name="focal_report.pdf", mime="application/pdf")

#with open("focal_report.xlsx", "rb") as f:
#    st.download_button("⬇️ Download Excel Report", f.read(), file_name="focal_report.xlsx")

with open("focal_report.html", "r") as f:
    st.download_button("⬇️ Download HTML Report", f.read(), file_name="focal_report.html")

import base64

def show_pdf_in_streamlit(pdf_path):
    with open(pdf_path, "rb") as f:
        base64_pdf = base64.b64encode(f.read()).decode("utf-8")
    pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="700" height="1000" type="application/pdf"></iframe>'
    st.markdown(pdf_display, unsafe_allow_html=True)

export_to_pdf(report_df)  # Your PDF generation function
show_pdf_in_streamlit("focal_report.pdf")



# 🌐 Global CMT Section
st.markdown(f"### 🌎 Peta Global CMT Harvard\n{cmt_start} – {cmt_end}")
def load_cmt(url):
    txt = requests.get(url).text
    lines = txt.split("\n")
    records = [lines[i:i+5] for i in range(0, len(lines), 5)]
    rows = []
    for rec in records:
        if len(rec) < 5: continue
        dt = f"{rec[0][5:15]} {rec[0][16:21]}"
        row = {
            'Datetime': dt,
            'Lat': float(rec[0][26:33]),
            'Lon': float(rec[0][35:41]),
            'Depth': float(rec[0][43:47]),
            'Mag_mb': float(rec[0][47:51]),
            'Mag_Ms': float(rec[0][52:55]),
            'S1': float(rec[4][56:60]),
            'D1': float(rec[4][61:64]),
            'R1': float(rec[4][65:69])
        }
        rows.append(row)
    return pd.DataFrame(rows)

from obspy.imaging.beachball import beach

def draw_beachballs(df, ax, projection, depth_col='Depth', lon_col='Lon', lat_col='Lat', scale=1.0):
    for _, row in df.iterrows():
        if all(pd.notnull(row[col]) for col in ['S1', 'D1', 'R1']):
            x, y = projection.transform_point(row[lon_col], row[lat_col], ccrs.Geodetic())
            color = "r" if row[depth_col] < 60 else "y" if row[depth_col] < 300 else "g"
            bb = beach([row['S1'], row['D1'], row['R1']],
                       xy=(x, y), width=scale,
                       linewidth=0.5, alpha=0.65,
                       zorder=10, facecolor=color)
            ax.add_collection(bb)


urls = [
    "https://www.ldeo.columbia.edu/~gcmt/projects/CMT/catalog/jan76_dec20.ndk",
    "https://www.ldeo.columbia.edu/~gcmt/projects/CMT/catalog/PRE1976/deep_1962-1976.ndk",
    "https://www.ldeo.columbia.edu/~gcmt/projects/CMT/catalog/PRE1976/intdep_1962-1975.ndk"
]

df_cmt = pd.concat([load_cmt(url) for url in urls])

df_cmt['Datetime'] = pd.to_datetime(df_cmt['Datetime'], errors='coerce')
df_cmt = df_cmt[
    (df_cmt['Datetime'] >= cmt_start) & (df_cmt['Datetime'] <= cmt_end) &
    (df_cmt['Lat'].between(South, North)) & (df_cmt['Lon'].between(West, East))
]

# 🗺️ Plot Global CMT
fig2 = plt.figure(dpi=300)
ax2 = fig2.add_subplot(111, projection=ccrs.PlateCarree(central_longitude=120))
ax2.set_extent((West, East, South - 0.5, North + 0.5))
ax2.add_feature(cfeature.BORDERS, linestyle='-', linewidth=0.5, alpha=0.5)
ax2.coastlines(resolution='10m', color='black', linewidth=0.5, alpha=0.5)
#scale_factor = compute_beachball_scale(West, East, South, North)
draw_beachballs(df_cmt, ax2, ax2.projection, depth_col='Depth', lon_col='Lon', lat_col='Lat', scale=w)

st.pyplot(fig2)
df_cmt.index = range(1, len(df_cmt) + 1)  # Reindex starting from 1
st.dataframe(df_cmt)
